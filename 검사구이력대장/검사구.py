# -*- coding: utf-8 -*-
import json
import calendar
import os
import shutil
import webbrowser
import base64
from datetime import datetime
from html import escape
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import load_workbook
from PIL import Image, ImageTk
import qrcode


APP_TITLE = "검사구 관리 대장"
APP_SUBTITLE = "검사구 관리번호, 점검이력, QR 코드, HTML 이력카드를 한 번에 관리하는 도구"
BASE_DIR = Path(__file__).resolve().parent
RUNTIME_DIR = BASE_DIR.parent / "inspection_tool_runtime"
DATA_DIR = RUNTIME_DIR / "data"
EXPORT_DIR = BASE_DIR / "export"
QR_DIR = EXPORT_DIR / "qrcode"
CARD_DIR = EXPORT_DIR / "cards"
PEOPLE_DIR = EXPORT_DIR / "people"
DB_PATH = DATA_DIR / "inspection_tools.json"
CONFIG_PATH = DATA_DIR / "inspection_config.json"

UI = {
    "bg": "#F3F6FB",
    "panel": "#FFFFFF",
    "soft": "#EDF3FB",
    "line": "#D6DEEB",
    "text": "#0F172A",
    "muted": "#64748B",
    "accent": "#2563EB",
    "accent_soft": "#1D4ED8",
    "secondary": "#EDF3FB",
    "danger": "#DC2626",
    "header_bg": "#0F172A",
    "header_panel": "#172036",
    "header_text": "#FFFFFF",
}

APP_BG = UI["bg"]
SURFACE_BG = UI["panel"]
SURFACE_ALT = "#EDF3FB"
HEADER_BG = UI["header_bg"]
HEADER_SUB = "#94A3B8"
PRIMARY = UI["accent"]
PRIMARY_ACTIVE = "#1D4ED8"
TEXT_MAIN = UI["text"]
TEXT_MUTED = UI["muted"]
BORDER = UI["line"]
TABLE_BG = "#F8FAFC"

PANEL_PAD = 10
SIDE_PANEL_WIDTH = 260
LEFT_PANEL_WIDTH = 445
PHOTO_BOX_WIDTH = 280
PHOTO_BOX_HEIGHT = 210
SIDE_PHOTO_WIDTH = 220
SIDE_PHOTO_HEIGHT = 180
QR_PREVIEW_SIZE = 72
FIELD_LABEL_WIDTH = 9
HEADER_LOGO_PATH = BASE_DIR / "logo.png"
PERSON_PHOTO_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif")

DEFAULT_CONFIG = {
    "workbook_path": "",
    "nas_sync_dir": "",
    "nas_base_url": "",
    "netlify_site_url": "",
    "webdav_url": "",
    "webdav_username": "",
    "webdav_password": "",
    "company_name": "품질팀",
    "scan_prefix": "TOOL:",
}


def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    QR_DIR.mkdir(parents=True, exist_ok=True)
    CARD_DIR.mkdir(parents=True, exist_ok=True)
    PEOPLE_DIR.mkdir(parents=True, exist_ok=True)


def load_config():
    ensure_dirs()
    if not CONFIG_PATH.exists():
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG.copy()
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        save_config(DEFAULT_CONFIG)
        return DEFAULT_CONFIG.copy()
    merged = DEFAULT_CONFIG.copy()
    merged.update(data)
    return merged


def save_config(config):
    CONFIG_PATH.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


def init_db():
    ensure_dirs()
    if not DB_PATH.exists():
        save_db({"tools": [], "inspections": []})


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def remove_light_edge_background(image):
    image = image.convert("RGBA")
    width, height = image.size
    if width <= 2 or height <= 2:
        return image

    pixels = image.load()
    corner_points = [
        pixels[0, 0],
        pixels[width - 1, 0],
        pixels[0, height - 1],
        pixels[width - 1, height - 1],
    ]
    bg = tuple(sum(point[index] for point in corner_points) // len(corner_points) for index in range(3))

    def is_background_pixel(x, y):
        r, g, b, a = pixels[x, y]
        if a == 0:
            return True
        distance = abs(r - bg[0]) + abs(g - bg[1]) + abs(b - bg[2])
        very_light = r >= 248 and g >= 248 and b >= 248
        return distance <= 56 or very_light

    visited = set()
    stack = []
    for x in range(width):
        stack.append((x, 0))
        stack.append((x, height - 1))
    for y in range(1, height - 1):
        stack.append((0, y))
        stack.append((width - 1, y))

    while stack:
        x, y = stack.pop()
        if (x, y) in visited or not is_background_pixel(x, y):
            continue
        visited.add((x, y))
        r, g, b, _a = pixels[x, y]
        pixels[x, y] = (r, g, b, 0)
        if x > 0:
            stack.append((x - 1, y))
        if x < width - 1:
            stack.append((x + 1, y))
        if y > 0:
            stack.append((x, y - 1))
        if y < height - 1:
            stack.append((x, y + 1))

    return image


def normalize_network_path(value):
    path_text = normalize_text(value)
    if not path_text:
        return ""
    path_text = path_text.replace("₩", "\\")
    if path_text.startswith("//"):
        return "\\\\" + path_text[2:].replace("/", "\\")
    return path_text


def load_db():
    init_db()
    try:
        return json.loads(DB_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = {"tools": [], "inspections": []}
        save_db(data)
        return data


def save_db(data):
    DB_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_plan_workbook(workbook_path):
    wb = load_workbook(workbook_path)
    records = {}
    for ws in wb.worksheets:
        if not normalize_text(ws.title).startswith("계획서"):
            continue
        for row in range(5, ws.max_row + 1):
            management_no = normalize_text(ws.cell(row=row, column=4).value)
            cycle = normalize_text(ws.cell(row=row, column=5).value)
            category = normalize_text(ws.cell(row=row, column=2).value)
            full_name = normalize_text(ws.cell(row=row, column=3).value)
            if not any([management_no, cycle, category, full_name]):
                continue
            if not management_no:
                continue
            quarter_marks = []
            for col in range(6, 10):
                marker = normalize_text(ws.cell(row=row, column=col).value)
                if marker:
                    quarter_marks.append(f"{col - 5}/4Q:{marker}")
            note = normalize_text(ws.cell(row=row, column=10).value)
            if quarter_marks:
                note = " / ".join(quarter_marks + ([note] if note else []))
            records[management_no] = {
                "management_no": management_no,
                "category": category,
                "item_name": category,
                "full_name": full_name,
                "cycle": cycle,
                "storage_location": "측정실",
                "department": "품질팀",
                "inspection_method": "체크시트",
                "specification": "",
                "notes": note,
            }
    return list(records.values())


def upsert_tools(records):
    created = 0
    updated = 0
    timestamp = now_text()
    data = load_db()
    tool_map = {item["management_no"]: item for item in data["tools"]}
    for record in records:
        existing = tool_map.get(record["management_no"])
        if existing:
            existing.update(
                {
                    "category": record["category"],
                    "item_name": record["item_name"],
                    "full_name": record["full_name"],
                    "cycle": record["cycle"],
                    "storage_location": record["storage_location"],
                    "department": record["department"],
                    "inspection_method": record["inspection_method"],
                    "specification": record["specification"],
                    "notes": record["notes"],
                    "updated_at": timestamp,
                }
            )
            updated += 1
        else:
            tool_map[record["management_no"]] = {
                "management_no": record["management_no"],
                "category": record["category"],
                "item_name": record["item_name"],
                "full_name": record["full_name"],
                "cycle": record["cycle"],
                "storage_location": record["storage_location"],
                "department": record["department"],
                "inspection_method": record["inspection_method"],
                "specification": record["specification"],
                "notes": record["notes"],
                "image_path": normalize_text(record.get("image_path")),
                "created_at": timestamp,
                "updated_at": timestamp,
            }
            created += 1
    data["tools"] = sorted(tool_map.values(), key=lambda item: item["management_no"])
    save_db(data)
    return created, updated


def list_tools(keyword=""):
    keyword = normalize_text(keyword)
    rows = load_db()["tools"]
    if keyword:
        rows = [
            row for row in rows
            if keyword.lower() in normalize_text(row.get("management_no")).lower()
            or keyword.lower() in normalize_text(row.get("customer")).lower()
            or keyword.lower() in normalize_text(row.get("category")).lower()
            or keyword.lower() in normalize_text(row.get("full_name")).lower()
        ]
    return sorted(rows, key=lambda item: item["management_no"])


def get_tool(management_no):
    for item in load_db()["tools"]:
        if item["management_no"] == management_no:
            return item
    return None


def save_tool_record(record):
    management_no = normalize_text(record.get("management_no"))
    if not management_no:
        raise ValueError("관리번호는 필수입니다.")
    timestamp = now_text()
    data = load_db()
    tools = data["tools"]
    for item in tools:
        if item["management_no"] == management_no:
            item.update(
                {
                    "customer": normalize_text(record.get("customer")),
                    "category": normalize_text(record.get("category")),
                    "item_name": normalize_text(record.get("item_name")),
                    "full_name": normalize_text(record.get("full_name")),
                    "cycle": normalize_text(record.get("cycle")),
                    "storage_location": normalize_text(record.get("storage_location")),
                    "department": normalize_text(record.get("department")),
                    "inspection_method": normalize_text(record.get("inspection_method")),
                    "specification": normalize_text(record.get("specification")),
                    "production_date": normalize_text(record.get("production_date")),
                    "maker": normalize_text(record.get("maker")),
                    "notes": normalize_text(record.get("notes")),
                    "image_path": normalize_text(record.get("image_path")),
                    "updated_at": timestamp,
                }
            )
            save_db(data)
            return
    tools.append(
        {
            "management_no": management_no,
            "customer": normalize_text(record.get("customer")),
            "category": normalize_text(record.get("category")),
            "item_name": normalize_text(record.get("item_name")),
            "full_name": normalize_text(record.get("full_name")),
            "cycle": normalize_text(record.get("cycle")),
            "storage_location": normalize_text(record.get("storage_location")),
            "department": normalize_text(record.get("department")),
            "inspection_method": normalize_text(record.get("inspection_method")),
            "specification": normalize_text(record.get("specification")),
            "production_date": normalize_text(record.get("production_date")),
            "maker": normalize_text(record.get("maker")),
            "notes": normalize_text(record.get("notes")),
            "image_path": normalize_text(record.get("image_path")),
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    )
    data["tools"] = sorted(tools, key=lambda item: item["management_no"])
    save_db(data)


def delete_tool_record(management_no):
    management_no = normalize_text(management_no)
    if not management_no:
        raise ValueError("관리번호를 먼저 선택하세요.")
    data = load_db()
    original_count = len(data["tools"])
    data["tools"] = [item for item in data["tools"] if item["management_no"] != management_no]
    data["inspections"] = [item for item in data["inspections"] if item["management_no"] != management_no]
    if len(data["tools"]) == original_count:
        raise ValueError("삭제할 검사구를 찾을 수 없습니다.")
    save_db(data)


def delete_tool_records(management_numbers):
    targets = {normalize_text(management_no) for management_no in management_numbers if normalize_text(management_no)}
    if not targets:
        raise ValueError("삭제할 검사구를 먼저 선택하세요.")
    data = load_db()
    original_count = len(data["tools"])
    data["tools"] = [item for item in data["tools"] if item["management_no"] not in targets]
    data["inspections"] = [item for item in data["inspections"] if item["management_no"] not in targets]
    deleted_count = original_count - len(data["tools"])
    if deleted_count == 0:
        raise ValueError("삭제할 검사구를 찾을 수 없습니다.")
    save_db(data)
    return deleted_count


def add_inspection_record(record):
    management_no = normalize_text(record.get("management_no"))
    if not management_no:
        raise ValueError("관리번호를 먼저 선택하세요.")
    data = load_db()
    data["inspections"].append(
        {
            "id": len(data["inspections"]) + 1,
            "management_no": management_no,
            "inspection_date": normalize_text(record.get("inspection_date")),
            "master_sample_match": normalize_text(record.get("master_sample_match")),
            "storage_status": normalize_text(record.get("storage_status")),
            "cleaning_status": normalize_text(record.get("cleaning_status")),
            "wear_status": normalize_text(record.get("wear_status")),
            "fit_status": normalize_text(record.get("fit_status")),
            "result_text": normalize_text(record.get("result_text")),
            "usage_flag": normalize_text(record.get("usage_flag")),
            "author": normalize_text(record.get("author")),
            "reviewer": normalize_text(record.get("reviewer")),
            "approver": normalize_text(record.get("approver")),
            "memo": normalize_text(record.get("memo")),
            "created_at": now_text(),
        }
    )
    save_db(data)


def update_inspection_record(inspection_id, record):
    data = load_db()
    for item in data["inspections"]:
        if int(item.get("id", 0)) == int(inspection_id):
            item.update(
                {
                    "inspection_date": normalize_text(record.get("inspection_date")),
                    "master_sample_match": normalize_text(record.get("master_sample_match")),
                    "storage_status": normalize_text(record.get("storage_status")),
                    "cleaning_status": normalize_text(record.get("cleaning_status")),
                    "wear_status": normalize_text(record.get("wear_status")),
                    "fit_status": normalize_text(record.get("fit_status")),
                    "result_text": normalize_text(record.get("result_text")),
                    "usage_flag": normalize_text(record.get("usage_flag")),
                    "author": normalize_text(record.get("author")),
                    "reviewer": normalize_text(record.get("reviewer")),
                    "approver": normalize_text(record.get("approver")),
                    "memo": normalize_text(record.get("memo")),
                    "updated_at": now_text(),
                }
            )
            save_db(data)
            return
    raise ValueError("수정할 점검 이력을 찾을 수 없습니다.")


def get_inspection_record(inspection_id):
    for item in load_db()["inspections"]:
        if int(item.get("id", 0)) == int(inspection_id):
            return item
    return None


def delete_inspection_record(inspection_id):
    data = load_db()
    original_count = len(data["inspections"])
    data["inspections"] = [item for item in data["inspections"] if int(item.get("id", 0)) != int(inspection_id)]
    if len(data["inspections"]) == original_count:
        raise ValueError("삭제할 점검 이력을 찾을 수 없습니다.")
    save_db(data)


def list_inspections(management_no):
    rows = [row for row in load_db()["inspections"] if row["management_no"] == management_no]
    rows.sort(key=lambda item: (normalize_text(item.get("inspection_date")) or normalize_text(item.get("created_at")), item.get("id", 0)))
    return rows


def public_base_url(config):
    return normalize_text(config.get("netlify_site_url")) or normalize_text(config.get("nas_base_url"))


def qr_payload_for_tool(tool, config):
    management_no = tool["management_no"]
    base_url = public_base_url(config)
    if base_url:
        base_url = base_url.rstrip("/")
        if base_url.lower().endswith("/cards"):
            return f"{base_url}/{management_no}.html"
        return f"{base_url}/cards/{management_no}.html"
    return f"{normalize_text(config.get('scan_prefix'))}{management_no}"


def is_quickconnect_url(url_text):
    lowered = normalize_text(url_text).lower()
    return "quickconnect.to/" in lowered or "quickconnect." in lowered


def render_qr_image(data, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image = qrcode.make(data)
    image.save(output_path)
    return output_path


def person_photo_path(person_name):
    person_name = normalize_text(person_name)
    if not person_name or person_name == "-":
        return None
    candidates = [
        PEOPLE_DIR / f"{person_name}{extension}"
        for extension in PERSON_PHOTO_EXTENSIONS
        if (PEOPLE_DIR / f"{person_name}{extension}").exists()
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def person_photo_src(person_name):
    photo_path = person_photo_path(person_name)
    if not photo_path:
        return ""
    version = int(photo_path.stat().st_mtime)
    return "../people/" + quote(photo_path.name) + f"?v={version}"


def person_html(person_name):
    person_name = normalize_text(person_name)
    if not person_name:
        return "-"
    src = person_photo_src(person_name)
    if not src:
        return escape(person_name)
    return (
        '<div class="person">'
        f'<img class="person-photo" src="{src}" alt="{escape(person_name)}">'
        f'<span>{escape(person_name)}</span>'
        '</div>'
    )


def person_profile_html(person_name):
    person_name = normalize_text(person_name)
    if not person_name:
        person_name = "담당자 미지정"
    src = person_photo_src(person_name)
    if src:
        photo_html = f'<img class="lead-photo" src="{src}" alt="{escape(person_name)}">'
    else:
        initial = escape(person_name[:1] or "?")
        photo_html = f'<div class="lead-photo lead-photo-empty">{initial}</div>'
    return (
        '<aside class="lead-person">'
        '<div class="lead-label">담당 검사자</div>'
        f'{photo_html}'
        f'<div class="lead-name">{escape(person_name)}</div>'
        '<div class="lead-note">최근 점검 담당자</div>'
        '</aside>'
    )


def sync_people_photos_to_nas(nas_web_root):
    people_files = [path for path in PEOPLE_DIR.iterdir() if path.is_file()]
    if not people_files:
        return
    nas_people_dir = nas_web_root / "people"
    nas_people_dir.mkdir(parents=True, exist_ok=True)
    for source in people_files:
        shutil.copy2(source, nas_people_dir / source.name)


def webdav_is_configured(config):
    return bool(normalize_text(config.get("webdav_url")))


def webdav_credentials(config):
    username = normalize_text(config.get("webdav_username"))
    password = normalize_text(config.get("webdav_password"))
    if not username or not password:
        raise ValueError("WebDAV 계정과 비밀번호를 설정해 주세요.")
    return username, password


def webdav_file_url(config, *parts):
    base_url = normalize_text(config.get("webdav_url")).rstrip("/")
    encoded_parts = [quote(str(part).strip("/")) for part in parts if str(part).strip("/")]
    if encoded_parts:
        return base_url + "/" + "/".join(encoded_parts)
    return base_url


def webdav_request(config, method, url, data=None, content_type="application/octet-stream"):
    username, password = webdav_credentials(config)
    token = base64.b64encode(f"{username}:{password}".encode("utf-8")).decode("ascii")
    headers = {"Authorization": f"Basic {token}"}
    if data is not None:
        headers["Content-Type"] = content_type
    request = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(request, timeout=30) as response:
            return response.status
    except HTTPError as exc:
        return exc.code
    except URLError as exc:
        raise ValueError(f"WebDAV 연결 실패: {exc}") from exc


def ensure_webdav_dirs(config):
    for dirname in ("cards", "qrcode", "people"):
        status = webdav_request(config, "MKCOL", webdav_file_url(config, dirname))
        if status not in (200, 201, 204, 405):
            raise ValueError(f"WebDAV 폴더 생성 실패: {dirname} (HTTP {status})")


def upload_file_to_webdav(config, source_path, *remote_parts):
    if not source_path.exists():
        return
    suffix = source_path.suffix.lower()
    content_type = "text/html; charset=utf-8" if suffix == ".html" else "image/png" if suffix == ".png" else "application/octet-stream"
    status = webdav_request(config, "PUT", webdav_file_url(config, *remote_parts), source_path.read_bytes(), content_type)
    if status not in (200, 201, 204):
        raise ValueError(f"WebDAV 업로드 실패: {source_path.name} (HTTP {status})")


def sync_people_photos_to_webdav(config):
    if not PEOPLE_DIR.exists():
        return
    for source in PEOPLE_DIR.iterdir():
        if source.is_file():
            upload_file_to_webdav(config, source, "people", source.name)


def upload_tool_assets_to_webdav(config, management_no):
    ensure_webdav_dirs(config)
    upload_file_to_webdav(config, QR_DIR / f"{management_no}.png", "qrcode", f"{management_no}.png")
    upload_file_to_webdav(config, CARD_DIR / f"{management_no}.html", "cards", f"{management_no}.html")
    sync_people_photos_to_webdav(config)


def upload_index_to_webdav(config, index_path):
    ensure_webdav_dirs(config)
    upload_file_to_webdav(config, index_path, "index.html")


def build_index_html(tools, config):
    base_url = public_base_url(config)
    rows = []
    for tool in sorted(tools, key=lambda item: normalize_text(item.get("management_no"))):
        management_no = normalize_text(tool.get("management_no"))
        if not management_no:
            continue
        card_href = f"cards/{escape(management_no)}.html"
        rows.append(
            f"""
            <tr>
              <td><a href="{card_href}">{escape(management_no)}</a></td>
              <td>{escape(normalize_text(tool.get("full_name")))}</td>
              <td>{escape(normalize_text(tool.get("category")))}</td>
              <td>{escape(normalize_text(tool.get("cycle")))}</td>
            </tr>
            """
        )
    body_rows = "\n".join(rows) if rows else '<tr><td colspan="4">생성된 이력카드가 없습니다.</td></tr>'
    public_url = escape(base_url.rstrip("/")) if base_url else "-"
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>검사구 이력카드 목록</title>
  <style>
    :root {{ --bg:#f5f7fb; --panel:#ffffff; --text:#10243e; --muted:#607089; --line:#d9e2f1; --accent:#2563eb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Arial,"Malgun Gothic",sans-serif; background:var(--bg); color:var(--text); }}
    main {{ max-width:1040px; margin:0 auto; padding:24px; }}
    h1 {{ margin:0 0 8px; font-size:28px; }}
    p {{ margin:0 0 18px; color:var(--muted); line-height:1.5; }}
    .panel {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:18px; }}
    table {{ width:100%; border-collapse:collapse; }}
    th,td {{ border-bottom:1px solid var(--line); padding:10px 8px; text-align:left; vertical-align:top; }}
    th {{ background:#f8fafc; }}
    a {{ color:var(--accent); font-weight:700; text-decoration:none; }}
    @media (max-width:720px) {{
      main {{ padding:16px; }}
      table {{ font-size:14px; }}
      th:nth-child(3),td:nth-child(3),th:nth-child(4),td:nth-child(4) {{ display:none; }}
    }}
  </style>
</head>
<body>
  <main>
    <h1>검사구 이력카드 목록</h1>
    <p>공개 주소: {public_url}<br>관리번호를 선택하면 해당 검사구 이력카드가 열립니다.</p>
    <section class="panel">
      <table>
        <thead><tr><th>관리번호</th><th>품명</th><th>차종</th><th>주기</th></tr></thead>
        <tbody>{body_rows}</tbody>
      </table>
    </section>
  </main>
</body>
</html>
"""


def export_index_page(config):
    index_html = build_index_html(list_tools(), config)
    index_path = EXPORT_DIR / "index.html"
    index_path.write_text(index_html, encoding="utf-8")
    nas_sync_dir = normalize_text(config.get("nas_sync_dir"))
    if nas_sync_dir:
        nas_paths = validate_nas_sync_dir(nas_sync_dir)
        try:
            (nas_paths["web_root"] / "index.html").write_text(index_html, encoding="utf-8")
        except PermissionError:
            (nas_paths["cards_dir"] / "index.html").write_text(index_html, encoding="utf-8")
    if webdav_is_configured(config):
        upload_index_to_webdav(config, index_path)
    return index_path


def resolve_nas_export_paths(nas_sync_dir):
    nas_root = Path(normalize_network_path(nas_sync_dir))
    if not nas_root:
        raise ValueError("NAS 동기화 폴더가 비어 있습니다.")
    if nas_root.name.lower() == "cards":
        web_root = nas_root.parent
        cards_dir = nas_root
    else:
        web_root = nas_root
        cards_dir = nas_root / "cards"
    return {
        "web_root": web_root,
        "cards_dir": cards_dir,
        "qrcode_dir": web_root / "qrcode",
    }


def ensure_nas_dirs(nas_sync_dir):
    paths = resolve_nas_export_paths(nas_sync_dir)
    for key in ("qrcode_dir", "cards_dir"):
        try:
            paths[key].mkdir(parents=True, exist_ok=True)
        except PermissionError:
            # Existing NAS folders can raise PermissionError on mkdir over SMB.
            # The later write test gives a clearer answer about actual access.
            pass
    return paths


def validate_nas_sync_dir(nas_sync_dir):
    try:
        paths = ensure_nas_dirs(nas_sync_dir)
        for key in ("cards_dir", "qrcode_dir"):
            probe_path = paths[key] / ".write_test.tmp"
            probe_path.write_text("ok", encoding="utf-8")
            probe_path.unlink(missing_ok=True)
        return paths
    except PermissionError as exc:
        raise ValueError(
            "NAS Web Station 폴더에 쓸 권한이 없습니다.\n"
            f"경로: {normalize_network_path(nas_sync_dir)}\n\n"
            "Windows 파일 탐색기에서 이 경로를 먼저 열어 NAS 계정으로 로그인하고, "
            "DSM에서 web 공유 폴더 쓰기 권한을 허용해 주세요."
        ) from exc
    except FileNotFoundError as exc:
        raise ValueError(
            "NAS Web Station 폴더 경로를 찾을 수 없습니다.\n"
            f"경로: {normalize_network_path(nas_sync_dir)}\n\n"
            "NAS IP와 공유 폴더 이름이 맞는지 확인해 주세요. 예: \\\\192.168.0.2\\web"
        ) from exc
    except OSError as exc:
        raise ValueError(
            "NAS Web Station 폴더에 연결할 수 없습니다.\n"
            f"경로: {normalize_network_path(nas_sync_dir)}\n"
            f"Windows 오류: {exc}\n\n"
            "파일 탐색기에서 \\\\192.168.0.2\\web 을 먼저 열어 접속되는지 확인해 주세요."
        ) from exc


def build_tool_html(tool, inspections, qr_payload):
    rows = []
    for index, item in enumerate(inspections, start=1):
        rows.append(
            f"""
            <tr>
              <td>{index}</td>
              <td>{escape(normalize_text(item['inspection_date']))}</td>
              <td>{escape(normalize_text(item['result_text']))}</td>
              <td>{escape(normalize_text(item['usage_flag']))}</td>
              <td>{person_html(item['author'])}</td>
              <td>{person_html(item['reviewer'])}</td>
              <td>{person_html(item['approver'])}</td>
              <td>{escape(normalize_text(item['memo']))}</td>
            </tr>
            """
        )
    history_html = "\n".join(rows) if rows else '<tr><td colspan="8">점검 이력이 없습니다.</td></tr>'
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(tool["management_no"])} 검사구 이력카드</title>
  <style>
    :root {{ --bg:#f5f7fb; --panel:#ffffff; --text:#10243e; --muted:#607089; --line:#d9e2f1; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:"Malgun Gothic",sans-serif; background:linear-gradient(180deg,#f5f7fb 0%,#edf5ef 100%); color:var(--text); }}
    .wrap {{ max-width:1120px; margin:0 auto; padding:24px; }}
    .hero,.table-card {{ background:var(--panel); border:1px solid var(--line); border-radius:24px; padding:24px; box-shadow:0 12px 28px rgba(16,36,62,.08); }}
    .title {{ font-size:30px; font-weight:700; margin:0 0 8px; }}
    .sub {{ color:var(--muted); }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:14px; }}
    .card {{ background:#f9fbff; border:1px solid var(--line); border-radius:18px; padding:16px; }}
    .label {{ color:var(--muted); font-size:13px; margin-bottom:6px; }}
    .value {{ font-size:18px; font-weight:700; word-break:break-word; }}
    .table-card {{ margin-top:22px; overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; min-width:860px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:12px 10px; text-align:left; vertical-align:top; }}
    th {{ background:#f3f7f4; }}
    .person {{ display:flex; align-items:center; gap:8px; min-width:110px; }}
    .person-photo {{ width:42px; height:42px; border-radius:8px; object-fit:cover; border:1px solid var(--line); background:#fff; }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1 class="title">검사구 이력 카드</h1>
      <p class="sub">QR 스캔 또는 관리번호 조회용 HTML 카드</p>
      <div class="grid">
        <div class="card"><div class="label">관리번호</div><div class="value">{escape(tool["management_no"])}</div></div>
        <div class="card"><div class="label">품목</div><div class="value">{escape(normalize_text(tool["full_name"]))}</div></div>
        <div class="card"><div class="label">차종 / 구분</div><div class="value">{escape(normalize_text(tool["category"]))}</div></div>
        <div class="card"><div class="label">점검주기</div><div class="value">{escape(normalize_text(tool["cycle"]))}</div></div>
        <div class="card"><div class="label">보관장소</div><div class="value">{escape(normalize_text(tool["storage_location"]))}</div></div>
        <div class="card"><div class="label">사용부서</div><div class="value">{escape(normalize_text(tool["department"]))}</div></div>
        <div class="card"><div class="label">점검방법</div><div class="value">{escape(normalize_text(tool["inspection_method"]))}</div></div>
        <div class="card"><div class="label">비고</div><div class="value">{escape(normalize_text(tool["notes"])) or "-"}</div></div>
      </div>
      <div class="qr">QR 연결 데이터: {escape(qr_payload)}</div>
      <div class="meta">최종 갱신: {escape(normalize_text(tool["updated_at"]))}</div>
    </section>
    <section class="table-card">
      <table>
        <thead>
          <tr><th>순번</th><th>일자</th><th>수리 및 점검결과</th><th>사용 유무</th><th>담당</th><th>승인</th><th>메모</th></tr>
        </thead>
        <tbody>{history_html}</tbody>
      </table>
    </section>
  </div>
</body>
</html>
"""


def build_tool_html_clean(tool, inspections, qr_payload):
    rows = []
    for index, item in enumerate(inspections, start=1):
        rows.append(
            f"""
            <tr>
              <td>{index}</td>
              <td>{escape(normalize_text(item['inspection_date']))}</td>
              <td>{escape(normalize_text(item['result_text']))}</td>
              <td>{escape(normalize_text(item['usage_flag']))}</td>
              <td>{person_html(item['author'])}</td>
              <td>{person_html(item['reviewer'])}</td>
              <td>{person_html(item['approver'])}</td>
              <td>{escape(normalize_text(item['memo']))}</td>
            </tr>
            """
        )
    history_html = "\n".join(rows) if rows else '<tr><td colspan="8">점검 이력이 없습니다.</td></tr>'
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(tool["management_no"])} 검사구 이력카드</title>
  <style>
    :root {{ --bg:#f5f7fb; --panel:#ffffff; --text:#10243e; --muted:#607089; --line:#d9e2f1; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:"Malgun Gothic",sans-serif; background:linear-gradient(180deg,#f5f7fb 0%,#edf5ef 100%); color:var(--text); }}
    .wrap {{ max-width:1120px; margin:0 auto; padding:24px; }}
    .hero,.table-card {{ background:var(--panel); border:1px solid var(--line); border-radius:24px; padding:24px; box-shadow:0 12px 28px rgba(16,36,62,.08); }}
    .title {{ font-size:30px; font-weight:700; margin:0 0 8px; }}
    .sub {{ color:var(--muted); }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:14px; }}
    .card {{ background:#f9fbff; border:1px solid var(--line); border-radius:18px; padding:16px; }}
    .label {{ color:var(--muted); font-size:13px; margin-bottom:6px; }}
    .value {{ font-size:18px; font-weight:700; word-break:break-word; }}
    .table-card {{ margin-top:22px; overflow:auto; }}
    table {{ width:100%; border-collapse:collapse; min-width:860px; }}
    th,td {{ border-bottom:1px solid var(--line); padding:12px 10px; text-align:left; vertical-align:top; }}
    th {{ background:#f3f7f4; }}
    .person {{ display:flex; align-items:center; gap:8px; min-width:110px; }}
    .person-photo {{ width:42px; height:42px; border-radius:8px; object-fit:cover; border:1px solid var(--line); background:#fff; }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1 class="title">검사구 이력 카드</h1>
      <p class="sub">QR 스캔 또는 관리번호 조회용 HTML 카드</p>
      <div class="grid">
        <div class="card"><div class="label">관리번호</div><div class="value">{escape(tool["management_no"])}</div></div>
        <div class="card"><div class="label">품명</div><div class="value">{escape(normalize_text(tool["full_name"]))}</div></div>
        <div class="card"><div class="label">차종 / 구분</div><div class="value">{escape(normalize_text(tool["category"]))}</div></div>
        <div class="card"><div class="label">점검주기</div><div class="value">{escape(normalize_text(tool["cycle"]))}</div></div>
        <div class="card"><div class="label">보관장소</div><div class="value">{escape(normalize_text(tool["storage_location"]))}</div></div>
        <div class="card"><div class="label">사용부서</div><div class="value">{escape(normalize_text(tool["department"]))}</div></div>
        <div class="card"><div class="label">점검방법</div><div class="value">{escape(normalize_text(tool["inspection_method"]))}</div></div>
        <div class="card"><div class="label">비고</div><div class="value">{escape(normalize_text(tool["notes"])) or "-"}</div></div>
      </div>
      <div class="qr">QR 연결 주소: {escape(qr_payload)}</div>
      <div class="meta">최종 갱신: {escape(normalize_text(tool["updated_at"]))}</div>
    </section>
    <section class="table-card">
      <table>
        <thead>
          <tr><th>순번</th><th>일자</th><th>점검결과</th><th>사용 유무</th><th>담당</th><th>승인</th><th>메모</th></tr>
        </thead>
        <tbody>{history_html}</tbody>
      </table>
    </section>
  </div>
</body>
</html>
"""


def build_tool_html_safe(tool, inspections, qr_payload):
    rows = []
    latest_inspection_date = normalize_text(inspections[0].get("inspection_date")) if inspections else "-"
    lead_person = normalize_text(inspections[0].get("reviewer")) if inspections else ""
    if not lead_person and inspections:
        lead_person = normalize_text(inspections[0].get("author"))
    lead_photo_src = person_photo_src(lead_person)
    if lead_photo_src:
        lead_photo_html = f'<img class="lead-photo" src="{lead_photo_src}" alt="{escape(lead_person)}">'
    else:
        lead_photo_html = '<div class="lead-photo lead-photo-empty">사진</div>'
    qr_img_src = "../qrcode/" + quote(f"{tool['management_no']}.png")
    for index, item in enumerate(inspections, start=1):
        reviewer_html = escape(normalize_text(item.get("reviewer")) or normalize_text(item.get("author")))
        approver_html = escape(normalize_text(item.get("approver")))
        approval_html = " / ".join(part for part in [reviewer_html, approver_html] if part) or "-"
        rows.append(
            f"""
            <tr>
              <td>{escape(normalize_text(item['inspection_date']))}</td>
              <td>{escape(normalize_text(item.get('master_sample_match'))) or "-"}</td>
              <td>{escape(normalize_text(item.get('storage_status')) or normalize_text(item.get('usage_flag'))) or "-"}</td>
              <td>{escape(normalize_text(item.get('cleaning_status'))) or "-"}</td>
              <td>{escape(normalize_text(item.get('wear_status'))) or "-"}</td>
              <td>{escape(normalize_text(item.get('fit_status'))) or "-"}</td>
              <td>{escape(normalize_text(item['result_text'])) or "-"}</td>
              <td>{approval_html}</td>
              <td>{escape(normalize_text(item['memo'])) or "-"}</td>
            </tr>
            """
        )
    history_html = "\n".join(rows) if rows else """
            <tr>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
              <td>&nbsp;</td>
            </tr>
            """
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(tool["management_no"])} 검사구 이력카드</title>
  <style>
    :root {{ --bg:#f5f7fb; --panel:#ffffff; --text:#111827; --muted:#5b6472; --line:#111827; --soft:#f3f4f6; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:"Malgun Gothic", Arial, sans-serif; background:#f5f7fb; color:var(--text); }}
    .wrap {{ max-width:1120px; margin:0 auto; padding:12px 18px; }}
    .sheet {{ position:relative; background:var(--panel); border:1px solid #d8dee8; border-radius:6px; padding:24px 34px 24px; box-shadow:0 8px 18px rgba(17,24,39,.06); }}
    .card-header {{ min-height:58px; padding-right:86px; margin-bottom:10px; }}
    .title {{ font-size:26px; font-weight:800; margin:0; letter-spacing:0; line-height:1.25; }}
    .sub {{ color:var(--muted); margin:0 0 18px; }}
    .top-grid {{ display:grid; grid-template-columns:140px 1fr; gap:8px; align-items:stretch; margin:10px 0 22px 0; }}
    .photo-card {{ background:#fff; border:1px solid #cfd8e6; display:flex; align-items:center; justify-content:center; overflow:hidden; min-height:222px; padding:5%; }}
    .qr-card {{ position:absolute; top:18px; right:34px; width:56px; height:56px; display:flex; align-items:center; justify-content:center; }}
    .lead-photo {{ width:100%; height:100%; object-fit:contain; display:block; }}
    .lead-photo-empty {{ width:100%; height:100%; display:flex; align-items:center; justify-content:center; font-size:20px; }}
    .qr-image {{ width:100%; height:100%; object-fit:contain; display:block; }}
    .info-table {{ width:100%; border-collapse:collapse; table-layout:fixed; border:1px solid #cfd8e6; }}
    .info-table th,.info-table td {{ border-bottom:1px solid #cfd8e6; border-right:1px solid #cfd8e6; padding:9px 10px; text-align:center; vertical-align:middle; }}
    .info-table tr:last-child th,.info-table tr:last-child td {{ border-bottom:0; }}
    .info-table th:last-child,.info-table td:last-child {{ border-right:0; }}
    .info-table th {{ width:13%; background:#f1f3f5; font-size:15px; font-weight:800; letter-spacing:0; white-space:nowrap; }}
    .info-table td {{ width:20%; font-size:15px; font-weight:700; word-break:break-word; }}
    .section-title {{ background:#f1f3f5; border:1px solid #cfd8e6; border-bottom:0; font-size:18px; font-weight:800; margin:22px 0 0; padding:8px; text-align:center; }}
    .table-card {{ overflow:auto; }}
    .history-table {{ width:100%; border-collapse:collapse; min-width:980px; border:1px solid #cfd8e6; }}
    .history-table th,.history-table td {{ border:1px solid #cfd8e6; padding:7px 6px; text-align:center; vertical-align:middle; }}
    .history-table th {{ background:#f1f3f5; font-size:15px; font-weight:800; line-height:1.3; }}
    .history-table td {{ font-size:13px; height:36px; }}
    @media (max-width:760px) {{
      .wrap {{ padding:12px; }}
      .sheet {{ padding:18px 14px; }}
      .top-grid {{ grid-template-columns:1fr; margin-right:0; }}
      .photo-card {{ width:140px; min-height:150px; }}
      .card-header {{ min-height:54px; padding-right:78px; }}
      .qr-card {{ width:62px; height:62px; right:14px; top:18px; }}
      .info-table th,.info-table td {{ padding:8px 6px; font-size:13px; }}
      .info-table th {{ letter-spacing:0; }}
      .history-table th {{ font-size:14px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="sheet">
      <div class="card-header">
        <h1 class="title">검사구 이력카드</h1>
      </div>
      <div class="top-grid">
      <div class="photo-card">{lead_photo_html}</div>
      <table class="info-table">
        <tbody>
          <tr>
            <th>관리번호</th>
            <td>{escape(normalize_text(tool["management_no"]))}</td>
            <th>고객사</th>
            <td>{escape(normalize_text(tool.get("customer"))) or "-"}</td>
          </tr>
          <tr>
            <th>차종</th>
            <td>{escape(normalize_text(tool["category"])) or "-"}</td>
            <th>품명</th>
            <td>{escape(normalize_text(tool["full_name"])) or "-"}</td>
          </tr>
          <tr>
            <th>품번</th>
            <td>{escape(normalize_text(tool["specification"])) or "-"}</td>
            <th>점검주기</th>
            <td>{escape(normalize_text(tool["cycle"])) or "-"}</td>
          </tr>
          <tr>
            <th>보관장소</th>
            <td>{escape(normalize_text(tool["storage_location"])) or "-"}</td>
            <th>사용부서</th>
            <td>{escape(normalize_text(tool["department"])) or "-"}</td>
          </tr>
          <tr>
            <th>제작일자</th>
            <td>{escape(normalize_text(tool.get("production_date"))) or "-"}</td>
            <th>제작업체</th>
            <td>{escape(normalize_text(tool.get("maker"))) or "-"}</td>
          </tr>
        </tbody>
      </table>
      </div>
      <div class="qr-card"><img class="qr-image" src="{qr_img_src}" alt="QR"></div>
      <div class="section-title">점검 결과</div>
      <div class="table-card">
      <table class="history-table">
        <thead>
          <tr>
            <th rowspan="2">점검일</th>
            <th rowspan="2">MASTER<br>SAMPLE<br>매칭상태</th>
            <th>보관상태</th>
            <th>청결 상태</th>
            <th>제품 매칭면</th>
            <th>제품안착시</th>
            <th rowspan="2">판정</th>
            <th rowspan="2">결재</th>
            <th rowspan="2">비고</th>
          </tr>
          <tr>
            <th>보관 여부</th>
            <th>청소여부</th>
            <th>마모상태</th>
            <th>유격상태</th>
          </tr>
        </thead>
        <tbody>{history_html}</tbody>
      </table>
      </div>
    </section>
  </div>
</body>
</html>
"""


def export_tool_assets(management_no, config, update_index=True, upload_webdav=True):
    tool = get_tool(management_no)
    if not tool:
        raise ValueError("관리번호를 찾을 수 없습니다.")
    inspections = list_inspections(management_no)
    qr_payload = qr_payload_for_tool(tool, config)
    qr_path = QR_DIR / f"{management_no}.png"
    card_path = CARD_DIR / f"{management_no}.html"
    render_qr_image(qr_payload, qr_path)
    card_path.write_text(build_tool_html_safe(tool, inspections, qr_payload), encoding="utf-8")

    nas_sync_dir = normalize_text(config.get("nas_sync_dir"))
    if nas_sync_dir:
        nas_paths = validate_nas_sync_dir(nas_sync_dir)
        sync_people_photos_to_nas(nas_paths["web_root"])
        render_qr_image(qr_payload, nas_paths["qrcode_dir"] / f"{management_no}.png")
        (nas_paths["cards_dir"] / f"{management_no}.html").write_text(build_tool_html_safe(tool, inspections, qr_payload), encoding="utf-8")
    if upload_webdav and webdav_is_configured(config):
        upload_tool_assets_to_webdav(config, management_no)
    if update_index:
        export_index_page(config)
    return qr_path, card_path, qr_payload


class ToolInspectionApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.configure_window_size()
        self.root.configure(bg=UI["bg"])
        self.config = load_config()
        self.selected_management_no = ""
        self.qr_preview_image = None
        self.qr_photo = None
        self.tool_image_photo = None
        self.tool_image_cache = {}
        self.header_logo_photo = None
        self.inspection_dialog = None
        self.inspection_memo = None
        self.editing_inspection_id = None
        self.init_inspection_vars()

        self.style = ttk.Style()
        self.build_styles()

        self.build_ui()
        self.auto_import_initial_workbook()
        self.refresh_tool_list()

    def configure_window_size(self):
        screen_width = max(self.root.winfo_screenwidth(), 1280)
        screen_height = max(self.root.winfo_screenheight(), 720)
        width = min(1760, screen_width - 40)
        height = min(980, screen_height - 70)
        self.root.geometry(f"{max(width, 1200)}x{max(height, 800)}")
        self.root.minsize(False, False)
        self.root.resizable(False, False)

    def build_styles(self):
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")
        self.style.configure(".", background=APP_BG, foreground=TEXT_MAIN)
        self.style.configure("App.TFrame", background=APP_BG)
        self.style.configure("Card.TFrame", background=SURFACE_BG, relief="flat")
        self.style.configure("Header.TFrame", background=HEADER_BG)
        self.style.configure("HeaderIcon.TLabel", background=HEADER_BG)
        self.style.configure("Title.TLabel", background=HEADER_BG, foreground="white", font=("Malgun Gothic", 24, "bold"))
        self.style.configure("Subtitle.TLabel", background=HEADER_BG, foreground=HEADER_SUB, font=("Malgun Gothic", 10))
        self.style.configure("Section.TLabel", background=SURFACE_BG, foreground=TEXT_MAIN, font=("Malgun Gothic", 12, "bold"))
        self.style.configure("Body.TLabel", background=SURFACE_BG, foreground=TEXT_MUTED, font=("Malgun Gothic", 10))
        self.style.configure("Status.TLabel", background=SURFACE_ALT, foreground=TEXT_MUTED, font=("Malgun Gothic", 10))
        self.style.configure(
            "Primary.TButton",
            background=PRIMARY,
            foreground="white",
            borderwidth=0,
            focusthickness=0,
            padding=(12, 8),
            font=("Malgun Gothic", 9, "bold"),
        )
        self.style.map(
            "Primary.TButton",
            background=[("active", PRIMARY_ACTIVE), ("pressed", PRIMARY_ACTIVE), ("disabled", "#E2E8F0")],
            foreground=[("active", "white"), ("pressed", "white"), ("disabled", "#64748B")],
        )
        self.style.configure(
            "Secondary.TButton",
            background=SURFACE_ALT,
            foreground=TEXT_MAIN,
            bordercolor=BORDER,
            lightcolor=SURFACE_ALT,
            darkcolor=SURFACE_ALT,
            padding=(12, 8),
            font=("Malgun Gothic", 9),
        )
        self.style.map(
            "Secondary.TButton",
            background=[("active", "#DCE7F6"), ("pressed", "#DCE7F6"), ("disabled", "#E2E8F0")],
            foreground=[("active", TEXT_MAIN), ("pressed", TEXT_MAIN), ("disabled", "#64748B")],
        )
        self.style.configure(
            "Danger.TButton",
            background=UI["danger"],
            foreground="white",
            borderwidth=0,
            focusthickness=0,
            padding=(12, 8),
            font=("Malgun Gothic", 9, "bold"),
        )
        self.style.map(
            "Danger.TButton",
            background=[("active", "#B91C1C"), ("pressed", "#B91C1C"), ("disabled", "#E2E8F0")],
            foreground=[("active", "white"), ("pressed", "white"), ("disabled", "#64748B")],
        )
        self.style.configure(
            "Modern.TEntry",
            fieldbackground=SURFACE_BG,
            foreground=TEXT_MAIN,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            padding=(9, 8),
        )
        self.style.configure(
            "Modern.Treeview",
            background=TABLE_BG,
            foreground=TEXT_MAIN,
            fieldbackground=TABLE_BG,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            rowheight=28,
            font=("Malgun Gothic", 9),
        )
        self.style.map("Modern.Treeview", background=[("selected", "#BFDBFE")], foreground=[("selected", TEXT_MAIN)])
        self.style.configure(
            "Modern.Treeview.Heading",
            background=SURFACE_BG,
            foreground=TEXT_MAIN,
            bordercolor=BORDER,
            lightcolor=BORDER,
            darkcolor=BORDER,
            padding=(8, 4),
            font=("Malgun Gothic", 9, "bold"),
        )
        self.style.map("Modern.Treeview.Heading", background=[("active", SURFACE_ALT)])

    def make_button(self, parent, text, command, variant="primary", padx=12):
        palettes = {
            "primary": (UI["accent"], "white", UI["accent_soft"], "white"),
            "secondary": (UI["secondary"], TEXT_MAIN, "#DCE7F6", TEXT_MAIN),
            "danger": (UI["danger"], "white", "#B91C1C", "white"),
        }
        bg, fg, active_bg, active_fg = palettes.get(variant, palettes["primary"])
        return tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=active_bg,
            activeforeground=active_fg,
            disabledforeground="#64748B",
            relief="flat",
            bd=0,
            highlightthickness=0,
            padx=padx,
            pady=5,
            font=("맑은 고딕", 9, "bold"),
            cursor="hand2",
        )

    def lock_tree_columns(self, tree):
        def block_resize(event):
            if tree.identify_region(event.x, event.y) == "separator":
                return "break"
            return None

        tree.bind("<Button-1>", block_resize, add="+")
        tree.bind("<B1-Motion>", block_resize, add="+")

    def refresh_header_logo(self):
        if not HEADER_LOGO_PATH.exists():
            self.header_logo_label.configure(image="", text="SEJI")
            return
        try:
            image = Image.open(HEADER_LOGO_PATH).convert("RGBA")
            max_width = 170
            max_height = 46
            ratio = min(max_width / max(image.width, 1), max_height / max(image.height, 1), 1)
            resized = image.resize(
                (max(1, int(image.width * ratio)), max(1, int(image.height * ratio))),
                Image.LANCZOS,
            )
            self.header_logo_photo = ImageTk.PhotoImage(resized)
            self.header_logo_label.configure(image=self.header_logo_photo, text="")
        except Exception:
            self.header_logo_label.configure(image="", text="SEJI")

    def _on_main_canvas_configure(self, event):
        if getattr(self, "main_canvas_window", None):
            self.main_canvas.itemconfigure(self.main_canvas_window, width=event.width)

    def _on_main_content_configure(self, _event):
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

    def _on_main_mousewheel(self, event):
        if event.delta:
            self.main_canvas.yview_scroll(int(-event.delta / 120), "units")

    def _bind_main_mousewheel(self, _event):
        self.main_canvas.bind_all("<MouseWheel>", self._on_main_mousewheel)

    def _unbind_main_mousewheel(self, _event):
        self.main_canvas.unbind_all("<MouseWheel>")

    def build_ui(self):
        self.root.configure(bg=APP_BG)
        header = ttk.Frame(self.root, style="Header.TFrame", padding=(28, 24, 28, 24))
        header.pack(fill="x", padx=0, pady=(0, 8))
        header.grid_columnconfigure(0, weight=1)
        title_wrap = ttk.Frame(header, style="Header.TFrame")
        title_wrap.grid(row=0, column=0, sticky="w")
        title_wrap.grid_columnconfigure(1, weight=1)
        self.header_logo_label = ttk.Label(title_wrap, style="HeaderIcon.TLabel")
        self.header_logo_label.grid(row=0, column=0, rowspan=2, sticky="w", padx=(0, 14))
        self.refresh_header_logo()
        ttk.Label(title_wrap, text=APP_TITLE, style="Title.TLabel").grid(row=0, column=1, sticky="w")
        ttk.Label(title_wrap, text=APP_SUBTITLE, style="Subtitle.TLabel").grid(row=1, column=1, sticky="w", pady=(6, 0))

        content_wrap = ttk.Frame(self.root, style="App.TFrame")
        content_wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        content_wrap.grid_rowconfigure(1, weight=1)
        content_wrap.grid_columnconfigure(0, weight=1)

        scan_panel = ttk.Frame(content_wrap, style="Card.TFrame", padding=(12, 10))
        scan_panel.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        scan_panel.grid_rowconfigure(0, minsize=36)
        scan_panel.grid_columnconfigure(1, weight=1)
        scan_panel.grid_columnconfigure(3, weight=1)
        ttk.Label(scan_panel, text="QR 리더 입력", style="Section.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.scan_var = tk.StringVar()
        scan_entry = ttk.Entry(scan_panel, textvariable=self.scan_var, style="Modern.TEntry", width=32)
        scan_entry.grid(row=0, column=1, sticky="ew", padx=(0, 8), ipady=1)
        scan_entry.bind("<Return>", self.on_scan_submit)
        button_row = ttk.Frame(scan_panel, style="Card.TFrame")
        button_row.grid(row=0, column=2, sticky="nsw")
        ttk.Button(button_row, text="조회", command=self.on_scan_submit, style="Primary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(button_row, text="엑셀 불러오기", command=self.import_workbook, style="Secondary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(button_row, text="전체 HTML/QR 갱신", command=self.export_all_assets, style="Primary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(button_row, text="설정", command=self.open_settings, style="Secondary.TButton").pack(side="left")
        self.status_var = tk.StringVar(value="준비 완료")
        ttk.Label(scan_panel, textvariable=self.status_var, style="Status.TLabel").grid(row=0, column=3, sticky="e", padx=(12, 0))

        body = ttk.Frame(content_wrap, style="App.TFrame")
        body.grid(row=1, column=0, sticky="nsew")
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=0, minsize=LEFT_PANEL_WIDTH)
        body.grid_columnconfigure(1, weight=1)

        left = ttk.Frame(body, style="Card.TFrame", padding=16, width=LEFT_PANEL_WIDTH)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left.grid_propagate(False)
        right = ttk.Frame(body, style="App.TFrame")
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(0, weight=48)
        right.grid_rowconfigure(1, weight=52)
        right.grid_columnconfigure(0, weight=1)
        right.grid_columnconfigure(1, weight=0, minsize=SIDE_PANEL_WIDTH)
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)

        search_row = ttk.Frame(left, style="Card.TFrame")
        search_row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_row, textvariable=self.search_var, style="Modern.TEntry", width=16)
        search_entry.pack(side="left", fill="x", expand=True)
        search_entry.bind("<KeyRelease>", lambda _event: self.refresh_tool_list())
        self.make_button(search_row, "추가", lambda: self.open_tool_edit_dialog(new_record=True), "primary", padx=9).pack(side="left", padx=(8, 0))
        self.make_button(search_row, "삭제", self.delete_selected_tool, "danger", padx=9).pack(side="left", padx=(6, 0))

        tool_tree_wrap = ttk.Frame(left, style="Card.TFrame")
        tool_tree_wrap.grid(row=1, column=0, sticky="nsew")
        tool_tree_wrap.grid_rowconfigure(0, weight=1)
        tool_tree_wrap.grid_columnconfigure(0, weight=1)
        tool_tree_wrap.grid_columnconfigure(1, minsize=16)
        self.tool_tree = ttk.Treeview(tool_tree_wrap, columns=("management_no", "category", "full_name", "cycle"), show="headings", height=12, selectmode="extended", style="Modern.Treeview")
        self.tool_tree.column("#0", width=0, minwidth=0, stretch=False)
        for name, title, width in [
            ("management_no", "관리번호", 80),
            ("category", "차종", 55),
            ("full_name", "품명", 205),
            ("cycle", "주기", 60),
        ]:
            self.tool_tree.heading(name, text=title, anchor="center")
            self.tool_tree.column(name, width=width, minwidth=width, anchor="w", stretch=False)
        tool_scroll_y = ttk.Scrollbar(tool_tree_wrap, orient="vertical", command=self.tool_tree.yview)
        self.tool_tree.configure(yscrollcommand=tool_scroll_y.set)
        self.tool_tree.grid(row=0, column=0, sticky="nsew")
        tool_scroll_y.grid(row=0, column=1, sticky="ns")
        self.lock_tree_columns(self.tool_tree)
        self.tool_tree.bind("<<TreeviewSelect>>", self.on_tool_select)
        self.tool_tree.bind("<Double-1>", lambda _event: self.open_tool_edit_dialog())

        self.form_panel = ttk.Frame(right, style="Card.TFrame", padding=10)
        self.form_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.qr_panel = ttk.Frame(right, style="Card.TFrame", padding=10, width=SIDE_PANEL_WIDTH)
        self.qr_panel.grid(row=0, column=1, sticky="nsew")
        self.qr_panel.grid_propagate(False)
        self.build_form_panel()
        self.build_qr_panel()

        self.history_panel = ttk.Frame(right, style="Card.TFrame", padding=10)
        self.history_panel.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(8, 0))
        self.build_history_panel()

    def build_form_panel(self):
        content = tk.Frame(self.form_panel, bg=UI["panel"])
        content.pack(fill="both", expand=True, padx=PANEL_PAD, pady=(0, 0))
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)
        content.grid_rowconfigure(0, weight=1)

        form_fields = tk.Frame(content, bg=UI["panel"])
        form_fields.grid(row=0, column=0, columnspan=2, sticky="nsew")

        self.tool_vars = {
            "management_no": tk.StringVar(),
            "customer": tk.StringVar(),
            "category": tk.StringVar(),
            "item_name": tk.StringVar(),
            "full_name": tk.StringVar(),
            "cycle": tk.StringVar(),
            "storage_location": tk.StringVar(),
            "department": tk.StringVar(),
            "inspection_method": tk.StringVar(),
            "specification": tk.StringVar(),
            "production_date": tk.StringVar(),
            "maker": tk.StringVar(),
            "image_path": tk.StringVar(),
        }
        self.notes_text = tk.Text(form_fields, width=1, height=2, font=("맑은 고딕", 10), wrap="word")
        self.tool_value_labels = {}

        form_fields.grid_columnconfigure(1, weight=1)
        form_fields.grid_columnconfigure(3, weight=1)
        for index, (key, label) in enumerate([
            ("management_no", "관리번호"),
            ("customer", "고객사"),
            ("category", "차종"),
            ("full_name", "품명"),
            ("specification", "품번"),
            ("cycle", "점검주기"),
            ("storage_location", "보관장소"),
            ("department", "사용부서"),
            ("production_date", "제작일"),
            ("maker", "제작처"),
        ]):
            row = index // 2
            label_col = 0 if index % 2 == 0 else 2
            value_col = label_col + 1
            tk.Label(
                form_fields,
                text=label,
                width=FIELD_LABEL_WIDTH,
                anchor="w",
                bg=UI["panel"],
                fg=UI["text"],
                font=("맑은 고딕", 10, "bold"),
            ).grid(row=row, column=label_col, sticky="w", padx=(0 if label_col == 0 else 18, 8), pady=5)
            value_box = tk.Frame(
                form_fields,
                bg="#F8FAFC",
                highlightthickness=1,
                highlightbackground=UI["line"],
            )
            value_box.grid(row=row, column=value_col, sticky="ew", padx=(0, 10), pady=5)
            value_label = tk.Label(
                value_box,
                textvariable=self.tool_vars[key],
                width=1,
                anchor="w",
                bg="#F8FAFC",
                fg=UI["text"],
                font=("맑은 고딕", 10),
            )
            value_label.pack(fill="x", padx=10, pady=6)
            self.tool_value_labels[key] = value_label

    def build_qr_panel(self):
        self.qr_preview_label = tk.Label(
            self.qr_panel,
            bg=UI["panel"],
            bd=0,
            highlightthickness=0,
        )
        # QR 위치 조절
        self.qr_preview_label.place(relx=1.0, x=-PANEL_PAD, y=4, anchor="ne")

        self.tool_photo_preview_width = SIDE_PHOTO_WIDTH
        self.tool_photo_preview_height = SIDE_PHOTO_HEIGHT
        photo_frame = tk.Frame(
            self.qr_panel,
            bg="#F8FAFC",
            highlightbackground=UI["line"],
            highlightthickness=1,
            width=SIDE_PHOTO_WIDTH,
            height=SIDE_PHOTO_HEIGHT,
        )
        # 검사구 가진 위치 조절
        photo_frame.pack(padx=PANEL_PAD, pady=(78, 12))
        photo_frame.pack_propagate(False)

        self.image_preview_label = tk.Label(
            photo_frame,
            text="사진 없음",
            bg="#F8FAFC",
            fg=UI["muted"],
            font=("맑은 고딕", 9),
            compound="center",
        )
        self.image_preview_label.pack(fill="both", expand=True)

        self.image_path_label = tk.Label(
            self.qr_panel,
            text="",
            bg=UI["panel"],
            fg=UI["muted"],
            font=("맑은 고딕", 8),
            wraplength=SIDE_PHOTO_WIDTH,
            justify="left",
        )
        self.image_path_label.pack_forget()
        self.qr_info_var = tk.StringVar(value="")
        self.qr_path_var = tk.StringVar(value="")

    def init_inspection_vars(self):
        self.inspection_vars = {
            "inspection_date": tk.StringVar(value=datetime.now().strftime("%Y-%m-%d")),
            "master_sample_match": tk.StringVar(),
            "storage_status": tk.StringVar(),
            "cleaning_status": tk.StringVar(),
            "wear_status": tk.StringVar(),
            "fit_status": tk.StringVar(),
            "result_text": tk.StringVar(),
            "usage_flag": tk.StringVar(value="사용"),
            "author": tk.StringVar(value=self.config.get("company_name", "품질팀")),
            "reviewer": tk.StringVar(),
            "approver": tk.StringVar(),
        }

    def build_history_panel(self):
        title_row = tk.Frame(self.history_panel, bg=UI["panel"])
        title_row.pack(fill="x", padx=PANEL_PAD, pady=(4, 4))
        tk.Label(title_row, text="점검 / 수리 이력", font=("맑은 고딕", 12, "bold"), bg=UI["panel"], fg=UI["text"]).pack(side="left")
        self.make_button(title_row, "이력 삭제", self.delete_selected_inspection, "danger", padx=10).pack(side="right", padx=(6, 0))       
        self.make_button(title_row, "이력 추가", self.open_inspection_dialog, "primary", padx=10).pack(side="right", padx=(6, 0))
        self.make_button(title_row, "이력 카드 열기", self.open_selected_card, "secondary", padx=10).pack(side="right")
        history_wrap = tk.Frame(self.history_panel, bg=UI["panel"])
        history_wrap.pack(fill="both", expand=True, padx=PANEL_PAD, pady=(0, 8))
        self.history_column_specs = [
            ("inspection_date", "점검일", 90),
            ("master_sample_match", "MASTER\nSAMPLE\n매칭상태", 118),
            ("storage_status", "보관 여부", 104),
            ("cleaning_status", "청소여부", 104),
            ("wear_status", "마모상태", 116),
            ("fit_status", "유격상태", 116),
            ("result_text", "판정", 84),
            ("approval", "결재", 84),
            ("memo", "비고", 96),
        ]
        self.build_history_header(history_wrap)
        self.history_tree = ttk.Treeview(
            history_wrap,
            columns=[column for column, _title, _width in self.history_column_specs],
            show="",
            height=7,
            style="Modern.Treeview",
        )
        self.history_tree.column("#0", width=0, minwidth=0, stretch=False)
        for name, _title, width in self.history_column_specs:
            self.history_tree.column(name, width=width, minwidth=width, anchor="center", stretch=False)
        self.history_tree.pack(fill="both", expand=True)
        self.lock_tree_columns(self.history_tree)
        self.history_tree.bind("<Double-1>", lambda _event: self.edit_selected_inspection())

    def build_history_header(self, parent):
        header_height = 64
        widths = [width for _name, _title, width in self.history_column_specs]
        total_width = sum(widths)
        canvas = tk.Canvas(parent, width=total_width, height=header_height, bg=UI["panel"], highlightthickness=0, bd=0)
        canvas.pack(anchor="w", fill="x")

        def draw_cell(x, y, width, height, text):
            canvas.create_rectangle(x, y, x + width, y + height, fill="#F1F3F5", outline="#CFD8E6")
            canvas.create_text(
                x + width / 2,
                y + height / 2,
                text=text,
                fill=UI["text"],
                font=("맑은 고딕", 8, "bold"),
                justify="center",
                width=max(width - 8, 40),
            )

        x_positions = []
        x = 0
        for width in widths:
            x_positions.append(x)
            x += width

        draw_cell(x_positions[0], 0, widths[0], header_height, "점검일")
        draw_cell(x_positions[1], 0, widths[1], header_height, "MASTER\nSAMPLE\n매칭상태")
        draw_cell(x_positions[2], 0, widths[2], 32, "보관상태")
        draw_cell(x_positions[3], 0, widths[3], 32, "청결 상태")
        draw_cell(x_positions[4], 0, widths[4], 32, "제품 매칭면")
        draw_cell(x_positions[5], 0, widths[5], 32, "제품안착시")
        draw_cell(x_positions[6], 0, widths[6], header_height, "판정")
        draw_cell(x_positions[7], 0, widths[7], header_height, "결재")
        draw_cell(x_positions[8], 0, widths[8], header_height, "비고")
        draw_cell(x_positions[2], 32, widths[2], 32, "보관 여부")
        draw_cell(x_positions[3], 32, widths[3], 32, "청소여부")
        draw_cell(x_positions[4], 32, widths[4], 32, "마모상태")
        draw_cell(x_positions[5], 32, widths[5], 32, "유격상태")

    def open_inspection_dialog(self, inspection=None):
        if not self.selected_management_no:
            messagebox.showwarning("검사구 선택", "왼쪽 목록에서 검사구를 먼저 선택하세요.")
            return
        if self.inspection_dialog is not None and self.inspection_dialog.winfo_exists():
            self.inspection_dialog.lift()
            self.inspection_dialog.focus_force()
            return
        self.inspection_dialog = tk.Toplevel(self.root)
        self.editing_inspection_id = int(inspection["id"]) if inspection else None
        self.inspection_dialog.title("이력 수정" if inspection else "이력 추가")
        self.inspection_dialog.configure(bg=UI["panel"])
        self.inspection_dialog.geometry("860x390")
        self.inspection_dialog.minsize(760, 360)
        self.inspection_dialog.resizable(True, True)
        self.inspection_dialog.transient(self.root)
        self.inspection_dialog.protocol("WM_DELETE_WINDOW", self.close_inspection_dialog)
        panel = tk.Frame(self.inspection_dialog, bg=UI["panel"])
        panel.pack(fill="both", expand=True, padx=18, pady=16)
        self.build_entry_panel(panel)
        if inspection:
            self.load_inspection_form(inspection)
        else:
            self.reset_inspection_form()
        self.inspection_dialog.grab_set()

    def close_inspection_dialog(self):
        if self.inspection_dialog is not None and self.inspection_dialog.winfo_exists():
            self.inspection_dialog.destroy()
        self.inspection_dialog = None
        self.inspection_memo = None
        self.editing_inspection_id = None

    def build_entry_panel(self, panel):
        panel.grid_columnconfigure(0, weight=1)

        top_btn_row = tk.Frame(panel, bg=UI["panel"])
        top_btn_row.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        save_text = "수정 저장" if self.editing_inspection_id else "이력 저장"
        self.make_button(top_btn_row, save_text, self.save_inspection, "primary", padx=14).pack(side="left")
        self.make_button(top_btn_row, "초기화", self.reset_inspection_form, "secondary", padx=12).pack(side="left", padx=(8, 0))

        person_btn_row = tk.Frame(top_btn_row, bg=UI["panel"])
        person_btn_row.pack(side="right")
        self.make_button(person_btn_row, "담당자 사진 등록", self.register_person_photo, "primary", padx=12).pack(side="left")
        self.make_button(person_btn_row, "사진 폴더", self.open_person_photo_folder, "secondary", padx=12).pack(side="left", padx=(8, 0))

        entry_card = tk.Frame(panel, bg=UI["soft"], highlightbackground=UI["line"], highlightthickness=1)
        entry_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        entry_card.grid_columnconfigure(1, weight=1)
        entry_card.grid_columnconfigure(3, weight=1)
        entry_card.grid_columnconfigure(5, weight=1)

        field_layout = [
            ("inspection_date", "점검일자", 0, 0),
            ("master_sample_match", "MASTER SAMPLE", 0, 2),
            ("storage_status", "보관 여부", 1, 0),
            ("cleaning_status", "청소여부", 1, 2),
            ("wear_status", "마모상태", 2, 0),
            ("fit_status", "유격상태", 2, 2),
            ("result_text", "판정", 3, 0),
            ("reviewer", "담당", 3, 2),
            ("approver", "승인", 3, 4),
        ]
        for key, label, row, label_col in field_layout:
            entry_col = label_col + 1
            label_width = 13 if key == "master_sample_match" else 9
            tk.Label(entry_card, text=label, width=label_width, anchor="w", bg=UI["soft"], fg=UI["text"], font=("맑은 고딕", 10, "bold")).grid(row=row, column=label_col, sticky="w", padx=(14 if label_col == 0 else 18, 8), pady=(12 if row == 0 else 8, 10 if row == 3 else 4))
            if key == "inspection_date":
                date_row = tk.Frame(entry_card, bg=UI["soft"])
                date_row.grid(row=row, column=entry_col, sticky="ew", padx=(0, 14), pady=(12 if row == 0 else 8, 10 if row == 3 else 4))
                date_row.grid_columnconfigure(0, weight=1)
                date_entry = tk.Entry(date_row, textvariable=self.inspection_vars[key], font=("맑은 고딕", 10), relief="solid", bd=1, state="readonly", cursor="hand2")
                date_entry.grid(row=0, column=0, sticky="ew", ipady=3)
                date_entry.bind("<Button-1>", lambda _event: self.open_inspection_date_picker())
            else:
                tk.Entry(entry_card, textvariable=self.inspection_vars[key], font=("맑은 고딕", 10), relief="solid", bd=1).grid(row=row, column=entry_col, sticky="ew", padx=(0, 14), pady=(12 if row == 0 else 8, 10 if row == 3 else 4), ipady=3)

        tk.Label(entry_card, text="비고", width=9, anchor="w", bg=UI["soft"], fg=UI["text"], font=("맑은 고딕", 10, "bold")).grid(row=4, column=0, sticky="nw", padx=(14, 8), pady=(4, 12))
        self.inspection_memo = tk.Text(entry_card, width=1, height=3, font=("맑은 고딕", 10), relief="solid", bd=1, wrap="word")
        self.inspection_memo.grid(row=4, column=1, columnspan=5, sticky="ew", padx=(0, 14), pady=(4, 12))

    def open_inspection_date_picker(self):
        current = self.inspection_vars["inspection_date"].get().strip()
        try:
            selected = datetime.strptime(current, "%Y-%m-%d")
        except ValueError:
            selected = datetime.now()

        picker = tk.Toplevel(self.inspection_dialog or self.root)
        picker.title("점검일자 선택")
        picker.configure(bg=UI["panel"])
        picker.resizable(False, False)
        picker.transient(self.inspection_dialog or self.root)
        picker.grab_set()

        state = {"year": selected.year, "month": selected.month}
        header = tk.Frame(picker, bg=UI["panel"])
        header.pack(fill="x", padx=12, pady=(12, 6))
        body = tk.Frame(picker, bg=UI["panel"])
        body.pack(padx=12, pady=(0, 12))
        title_var = tk.StringVar()

        def choose(day):
            self.inspection_vars["inspection_date"].set(f"{state['year']:04d}-{state['month']:02d}-{day:02d}")
            picker.destroy()

        def draw_month():
            for child in body.winfo_children():
                child.destroy()
            title_var.set(f"{state['year']:04d}년 {state['month']:02d}월")
            for col, day_name in enumerate(["월", "화", "수", "목", "금", "토", "일"]):
                tk.Label(body, text=day_name, width=4, bg=UI["panel"], fg=UI["muted"], font=("맑은 고딕", 9, "bold")).grid(row=0, column=col, pady=(0, 4))
            for row_index, week in enumerate(calendar.monthcalendar(state["year"], state["month"]), start=1):
                for col, day in enumerate(week):
                    if day == 0:
                        tk.Label(body, text="", width=4, bg=UI["panel"]).grid(row=row_index, column=col, padx=2, pady=2)
                        continue
                    tk.Button(
                        body,
                        text=str(day),
                        width=4,
                        relief="flat",
                        bg=UI["soft"],
                        fg=UI["text"],
                        activebackground="#DBEAFE",
                        command=lambda value=day: choose(value),
                    ).grid(row=row_index, column=col, padx=2, pady=2)

        def move_month(delta):
            month = state["month"] + delta
            if month < 1:
                state["year"] -= 1
                month = 12
            elif month > 12:
                state["year"] += 1
                month = 1
            state["month"] = month
            draw_month()

        self.make_button(header, "<", lambda: move_month(-1), "secondary", padx=8).pack(side="left")
        tk.Label(header, textvariable=title_var, width=16, bg=UI["panel"], fg=UI["text"], font=("맑은 고딕", 11, "bold")).pack(side="left", expand=True)
        self.make_button(header, ">", lambda: move_month(1), "secondary", padx=8).pack(side="right")
        draw_month()

    def auto_import_initial_workbook(self):
        if list_tools():
            return
        desktop_dir = Path(os.environ.get("USERPROFILE", str(Path.home()))) / "Desktop"
        candidates = [desktop_dir / name for name in os.listdir(desktop_dir) if name.lower().endswith(".xlsx")]
        if candidates:
            try:
                records = parse_plan_workbook(candidates[0])
                created, updated = upsert_tools(records)
                self.config["workbook_path"] = str(candidates[0])
                save_config(self.config)
                self.status_var.set(f"초기 엑셀 불러오기 완료: 신규 {created}건 / 갱신 {updated}건")
            except Exception as exc:
                self.status_var.set(f"초기 엑셀 불러오기 실패: {exc}")

    def refresh_tool_list(self):
        for item in self.tool_tree.get_children():
            self.tool_tree.delete(item)
        for row in list_tools(self.search_var.get()):
            self.tool_tree.insert("", "end", iid=row["management_no"], values=(row["management_no"], row["category"], row["full_name"], row["cycle"]))

    def refresh_tool_image_preview(self, image_path=""):
        if not hasattr(self, "image_preview_label"):
            return
        path_text = normalize_text(image_path)
        if not path_text:
            self.tool_image_photo = None
            self.image_preview_label.configure(image="", text="사진 없음")
            self.image_path_label.configure(text="")
            return
        image_file = Path(path_text)
        if not image_file.exists():
            self.tool_image_photo = None
            self.image_preview_label.configure(image="", text="사진 경로 없음")
            self.image_path_label.configure(text=path_text)
            return
        try:
            preview_width = getattr(self, "tool_photo_preview_width", PHOTO_BOX_WIDTH)
            preview_height = getattr(self, "tool_photo_preview_height", PHOTO_BOX_HEIGHT)
            stat = image_file.stat()
            cache_key = (str(image_file.resolve()), stat.st_mtime, stat.st_size, preview_width, preview_height)
            preview = self.tool_image_cache.get(cache_key)
            if preview is None:
                image = Image.open(image_file)
                image.thumbnail((preview_width - 16, preview_height - 16), Image.LANCZOS)
                image = remove_light_edge_background(image)
                preview = Image.new("RGBA", (preview_width, preview_height), "#F8FAFC")
                offset_x = (preview_width - image.width) // 2
                offset_y = (preview_height - image.height) // 2
                preview.paste(image, (offset_x, offset_y), image)
                self.tool_image_cache.clear()
                self.tool_image_cache[cache_key] = preview
            self.tool_image_photo = ImageTk.PhotoImage(preview)
            self.image_preview_label.configure(image=self.tool_image_photo, text="")
            self.image_path_label.configure(text=path_text)
        except Exception as exc:
            self.tool_image_photo = None
            self.image_preview_label.configure(image="", text="사진 로드 실패")
            self.image_path_label.configure(text=path_text)
            self.status_var.set(f"사진 로드 실패: {exc}")

    def choose_tool_image(self):
        if not self.tool_vars["management_no"].get().strip():
            messagebox.showwarning("검사구 선택", "사진을 넣을 검사구를 먼저 선택하거나 관리번호를 입력하세요.")
            return
        path = filedialog.askopenfilename(
            title="검사구 사진 선택",
            filetypes=[("Image", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("All Files", "*.*")],
        )
        if not path:
            return
        self.tool_vars["image_path"].set(path)
        self.refresh_tool_image_preview(path)
        self.save_tool_image_change("검사구 사진을 저장했습니다.")

    def clear_tool_image(self):
        self.tool_vars["image_path"].set("")
        self.refresh_tool_image_preview("")
        self.save_tool_image_change("검사구 사진을 지웠습니다.")

    def save_tool_image_change(self, status_text):
        try:
            save_tool_record(self.collect_tool_form())
            self.refresh_tool_list()
            management_no = self.tool_vars["management_no"].get().strip()
            if management_no:
                self.selected_management_no = management_no
                self.tool_tree.selection_set(management_no)
            self.status_var.set(status_text)
        except Exception as exc:
            messagebox.showerror("사진 저장 실패", str(exc))

    def open_tool_edit_dialog(self, new_record=False):
        if not new_record and not self.selected_management_no:
            messagebox.showwarning("검사구 선택", "수정할 검사구를 먼저 선택하세요.")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("새 검사구 등록" if new_record else "검사구 기본정보 수정")
        dialog.configure(bg=UI["panel"])
        dialog.geometry("560x585")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        panel = tk.Frame(dialog, bg=UI["panel"])
        panel.pack(fill="both", expand=True, padx=18, pady=16)
        panel.grid_columnconfigure(1, weight=1)

        fields = [
            ("management_no", "관리번호"),
            ("customer", "고객사"),
            ("category", "차종"),
            ("full_name", "품명"),
            ("specification", "품번"),
            ("cycle", "점검주기"),
            ("storage_location", "보관장소"),
            ("department", "사용부서"),
            ("production_date", "제작일"),
            ("maker", "제작처"),
        ]

        edit_vars = {}
        for row, (key, label) in enumerate(fields):
            value = "" if new_record else self.tool_vars[key].get()
            edit_vars[key] = tk.StringVar(value=value)
            tk.Label(
                panel,
                text=label,
                width=10,
                anchor="w",
                bg=UI["panel"],
                fg=UI["text"],
                font=("맑은 고딕", 10, "bold"),
            ).grid(row=row, column=0, sticky="w", padx=(0, 10), pady=5)
            tk.Entry(
                panel,
                textvariable=edit_vars[key],
                font=("맑은 고딕", 10),
                bg="#F8FAFC",
                fg=UI["text"],
                relief="solid",
                bd=1,
                highlightthickness=1,
                highlightbackground=UI["line"],
                highlightcolor=UI["accent"],
            ).grid(row=row, column=1, sticky="ew", pady=5, ipady=4)

        image_path_var = tk.StringVar(value="" if new_record else self.tool_vars["image_path"].get())
        image_row = tk.Frame(panel, bg=UI["panel"])
        image_row.grid(row=len(fields), column=0, columnspan=2, sticky="ew", pady=(8, 0))
        image_row.grid_columnconfigure(1, weight=1)
        tk.Label(
            image_row,
            text="사진",
            width=10,
            anchor="w",
            bg=UI["panel"],
            fg=UI["text"],
            font=("맑은 고딕", 10, "bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 10))
        image_path_label = tk.Label(
            image_row,
            textvariable=image_path_var,
            anchor="w",
            bg="#F8FAFC",
            fg=UI["muted"],
            font=("맑은 고딕", 9),
            relief="solid",
            bd=1,
        )
        image_path_label.grid(row=0, column=1, sticky="ew", ipady=5)

        def select_image():
            path = filedialog.askopenfilename(
                title="검사구 사진 선택",
                filetypes=[("Image", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("All Files", "*.*")],
            )
            if path:
                image_path_var.set(path)

        image_buttons = tk.Frame(image_row, bg=UI["panel"])
        image_buttons.grid(row=1, column=1, sticky="w", pady=(8, 0))
        self.make_button(image_buttons, "사진 선택", select_image, "secondary", padx=12).pack(side="left", padx=(0, 6))
        self.make_button(image_buttons, "사진 지우기", lambda: image_path_var.set(""), "secondary", padx=12).pack(side="left")

        button_row = tk.Frame(panel, bg=UI["panel"])
        button_row.grid(row=len(fields) + 1, column=0, columnspan=2, sticky="e", pady=(16, 0))

        def save_and_close():
            payload = {key: var.get() for key, var in edit_vars.items()}
            existing_tool = get_tool(self.selected_management_no) if self.selected_management_no else {}
            payload["item_name"] = normalize_text(existing_tool.get("item_name") or self.tool_vars["item_name"].get())
            payload["inspection_method"] = normalize_text(existing_tool.get("inspection_method") or self.tool_vars["inspection_method"].get())
            payload["image_path"] = image_path_var.get()
            payload["notes"] = normalize_text(existing_tool.get("notes") or self.notes_text.get("1.0", "end").strip())
            if not payload["management_no"].strip():
                messagebox.showwarning("관리번호", "관리번호는 필수입니다.")
                return
            try:
                save_tool_record(payload)
                self.refresh_tool_list()
                self.selected_management_no = payload["management_no"].strip()
                for key in self.tool_vars:
                    if key in payload:
                        self.tool_vars[key].set(normalize_text(payload.get(key)))
                self.tool_tree.selection_set(self.selected_management_no)
                self.tool_tree.see(self.selected_management_no)
                self.status_var.set("검사구 기본정보를 저장했습니다.")
                dialog.destroy()

                def refresh_saved_tool():
                    self.load_tool(self.selected_management_no)
                    self.refresh_tool_image_preview(payload["image_path"])
                    self.root.update_idletasks()

                self.root.after(50, refresh_saved_tool)
            except Exception as exc:
                messagebox.showerror("저장 실패", str(exc))

        self.make_button(button_row, "저장", save_and_close, "primary", padx=18).pack(side="left", padx=(0, 8))
        self.make_button(button_row, "닫기", dialog.destroy, "secondary", padx=16).pack(side="left")

    def on_tool_select(self, _event=None):
        selected = self.tool_tree.selection()
        if selected and selected[0] != self.selected_management_no:
            self.load_tool(selected[0])

    def load_tool(self, management_no):
        tool = get_tool(management_no)
        if not tool:
            return
        self.selected_management_no = management_no
        for key in self.tool_vars:
            self.tool_vars[key].set(normalize_text(tool.get(key, "")))
        self.notes_text.delete("1.0", "end")
        self.notes_text.insert("1.0", normalize_text(tool.get("notes", "")))
        self.refresh_tool_image_preview(tool.get("image_path", ""))
        self.refresh_history(management_no)
        self.refresh_qr_preview(management_no, tool)
        self.status_var.set(f"{management_no} 이력카드를 불러왔습니다.")

    def refresh_history(self, management_no):
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        for index, row in enumerate(list_inspections(management_no), start=1):
            reviewer = normalize_text(row.get("reviewer")) or normalize_text(row.get("author"))
            approver = normalize_text(row.get("approver"))
            approval = " / ".join(part for part in [reviewer, approver] if part)
            self.history_tree.insert(
                "",
                "end",
                iid=f"history-{row['id']}",
                values=(
                    row["inspection_date"],
                    normalize_text(row.get("master_sample_match")),
                    normalize_text(row.get("storage_status")) or normalize_text(row.get("usage_flag")),
                    normalize_text(row.get("cleaning_status")),
                    normalize_text(row.get("wear_status")),
                    normalize_text(row.get("fit_status")),
                    normalize_text(row.get("result_text")),
                    approval,
                    normalize_text(row.get("memo")),
                ),
            )

    def refresh_qr_preview(self, management_no, tool=None):
        try:
            if tool is None:
                tool = get_tool(management_no)
            if not tool:
                return
            qr_payload = qr_payload_for_tool(tool, self.config)
            if hasattr(self, "qr_preview_label"):
                self.qr_preview_image = qrcode.make(qr_payload).resize((QR_PREVIEW_SIZE, QR_PREVIEW_SIZE))
                self.qr_photo = ImageTk.PhotoImage(self.qr_preview_image)
                self.qr_preview_label.configure(image=self.qr_photo)
            if hasattr(self, "qr_info_var"):
                self.qr_info_var.set("")
            if hasattr(self, "qr_path_var"):
                self.qr_path_var.set("")
        except Exception as exc:
            if hasattr(self, "qr_preview_label"):
                self.qr_preview_label.configure(image="")
            if hasattr(self, "qr_info_var"):
                self.qr_info_var.set("")
            if hasattr(self, "qr_path_var"):
                self.qr_path_var.set("")
            self.status_var.set(f"QR 생성 실패: {exc}")

    def collect_tool_form(self):
        record = {key: self.tool_vars[key].get() for key in self.tool_vars}
        record["notes"] = self.notes_text.get("1.0", "end").strip()
        return record

    def save_current_tool(self):
        try:
            save_tool_record(self.collect_tool_form())
            self.refresh_tool_list()
            management_no = self.tool_vars["management_no"].get().strip()
            if management_no:
                self.selected_management_no = management_no
                self.load_tool(management_no)
            self.status_var.set("기본정보를 저장했습니다.")
        except Exception as exc:
            messagebox.showerror("저장 실패", str(exc))

    def save_inspection(self):
        if not self.selected_management_no:
            messagebox.showwarning("검사구 선택", "왼쪽 목록에서 검사구를 먼저 선택하세요.")
            return
        payload = {key: self.inspection_vars[key].get() for key in self.inspection_vars}
        payload["author"] = payload.get("reviewer", "")
        payload["usage_flag"] = payload.get("storage_status", "")
        payload["management_no"] = self.selected_management_no
        payload["memo"] = self.inspection_memo.get("1.0", "end").strip() if self.inspection_memo is not None else ""
        try:
            if self.editing_inspection_id:
                update_inspection_record(self.editing_inspection_id, payload)
                status_text = f"{self.selected_management_no} 점검 이력을 수정했습니다."
            else:
                add_inspection_record(payload)
                status_text = f"{self.selected_management_no} 점검 이력을 저장했습니다."
            self.refresh_history(self.selected_management_no)
            self.refresh_qr_preview(self.selected_management_no)
            self.reset_inspection_form()
            self.close_inspection_dialog()
            self.status_var.set(status_text)
        except Exception as exc:
            messagebox.showerror("이력 저장 실패", str(exc))

    def load_inspection_form(self, inspection):
        for key in self.inspection_vars:
            self.inspection_vars[key].set(normalize_text(inspection.get(key, "")))
        if not self.inspection_vars["storage_status"].get():
            self.inspection_vars["storage_status"].set(normalize_text(inspection.get("usage_flag", "")))
        if not self.inspection_vars["author"].get():
            self.inspection_vars["author"].set(normalize_text(inspection.get("reviewer", "")))
        if self.inspection_memo is not None and self.inspection_memo.winfo_exists():
            self.inspection_memo.delete("1.0", "end")
            self.inspection_memo.insert("1.0", normalize_text(inspection.get("memo", "")))

    def reset_inspection_form(self):
        self.inspection_vars["inspection_date"].set(datetime.now().strftime("%Y-%m-%d"))
        self.inspection_vars["master_sample_match"].set("")
        self.inspection_vars["storage_status"].set("")
        self.inspection_vars["cleaning_status"].set("")
        self.inspection_vars["wear_status"].set("")
        self.inspection_vars["fit_status"].set("")
        self.inspection_vars["result_text"].set("")
        self.inspection_vars["usage_flag"].set("")
        self.inspection_vars["author"].set("")
        self.inspection_vars["reviewer"].set("")
        self.inspection_vars["approver"].set("")
        if self.inspection_memo is not None and self.inspection_memo.winfo_exists():
            self.inspection_memo.delete("1.0", "end")

    def open_person_photo_folder(self):
        ensure_dirs()
        try:
            os.startfile(str(PEOPLE_DIR))
        except Exception as exc:
            messagebox.showerror("폴더 열기 실패", str(exc))

    def register_person_photo(self):
        ensure_dirs()
        default_name = normalize_text(self.inspection_vars["reviewer"].get())
        person_name = simpledialog.askstring(
            "담당자 사진 등록",
            "사진에 연결할 담당자 이름을 입력하세요.\n담당/승인 칸의 이름과 같아야 카드에 표시됩니다.",
            initialvalue=default_name,
            parent=self.root,
        )
        person_name = normalize_text(person_name)
        if not person_name:
            return
        source = filedialog.askopenfilename(
            title=f"{person_name} 담당자 사진 선택",
            filetypes=[("Image", "*.png;*.jpg;*.jpeg;*.webp;*.bmp;*.gif"), ("All Files", "*.*")],
        )
        if not source:
            return
        source_path = Path(source)
        if source_path.suffix.lower() not in PERSON_PHOTO_EXTENSIONS:
            messagebox.showwarning("사진 형식 확인", "jpg, png, webp, bmp, gif 파일을 선택해 주세요.")
            return
        target_path = PEOPLE_DIR / f"{person_name}{source_path.suffix.lower()}"
        try:
            shutil.copy2(source_path, target_path)
            self.status_var.set(f"{person_name} 담당자 사진을 등록했습니다.")
            messagebox.showinfo(
                "등록 완료",
                f"{person_name} 사진을 등록했습니다.\n선택 QR 생성 또는 전체 HTML/QR 갱신을 누르면 카드에 반영됩니다.",
            )
        except Exception as exc:
            messagebox.showerror("사진 등록 실패", str(exc))

    def prepare_new_tool(self):
        self.selected_management_no = ""
        for var in self.tool_vars.values():
            var.set("")
        self.notes_text.delete("1.0", "end")
        self.refresh_tool_image_preview("")
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        if hasattr(self, "qr_preview_label"):
            self.qr_preview_label.configure(image="")
        if hasattr(self, "qr_info_var"):
            self.qr_info_var.set("새 검사구를 입력한 뒤 저장하면 QR과 HTML 카드가 생성됩니다.")
        if hasattr(self, "qr_path_var"):
            self.qr_path_var.set("")
        self.status_var.set("새 검사구 입력 모드")

    def _delete_selected_tool_legacy(self):
        selected_management_numbers = list(self.tool_tree.selection())
        if not selected_management_numbers and self.selected_management_no:
            selected_management_numbers = [self.selected_management_no]
        if not selected_management_numbers:
            messagebox.showwarning("검사구 선택", "삭제할 검사구를 먼저 선택하세요.")
            return
        selected_count = len(selected_management_numbers)
        if selected_count == 1:
            confirm_message = f"{selected_management_numbers[0]} 검사구와 연결된 점검 이력을 모두 삭제할까요?"
        else:
            preview = ", ".join(selected_management_numbers[:5])
            if selected_count > 5:
                preview += f" 외 {selected_count - 5}건"
            confirm_message = f"선택한 검사구 {selected_count}건을 삭제할까요?\n\n{preview}\n\n연결된 점검 이력도 모두 삭제됩니다."
        answer = messagebox.askyesno(
            "검사구 삭제",
            f"{self.selected_management_no} 검사구와 연결된 점검 이력을 모두 삭제할까요?",
        )
        if not answer:
            return
        try:
            deleted_count = delete_tool_records(selected_management_numbers)
            self.prepare_new_tool()
            self.search_var.set("")
            self.refresh_tool_list()
            self.status_var.set("검사구를 삭제했습니다.")
        except Exception as exc:
            messagebox.showerror("삭제 실패", str(exc))

    def delete_selected_tool(self):
        selected_management_numbers = list(self.tool_tree.selection())
        if not selected_management_numbers and self.selected_management_no:
            selected_management_numbers = [self.selected_management_no]
        if not selected_management_numbers:
            messagebox.showwarning("검사구 선택", "삭제할 검사구를 먼저 선택하세요.")
            return

        selected_count = len(selected_management_numbers)
        if selected_count == 1:
            confirm_message = f"{selected_management_numbers[0]} 검사구와 연결된 점검 이력을 모두 삭제할까요?"
        else:
            preview = ", ".join(selected_management_numbers[:5])
            if selected_count > 5:
                preview += f" 외 {selected_count - 5}건"
            confirm_message = f"선택한 검사구 {selected_count}건을 삭제할까요?\n\n{preview}\n\n연결된 점검 이력도 모두 삭제됩니다."

        if not messagebox.askyesno("검사구 삭제", confirm_message):
            return

        try:
            deleted_count = delete_tool_records(selected_management_numbers)
            self.prepare_new_tool()
            self.search_var.set("")
            self.refresh_tool_list()
            self.status_var.set(f"검사구 {deleted_count}건을 삭제했습니다.")
        except Exception as exc:
            messagebox.showerror("삭제 실패", str(exc))

    def delete_selected_inspection(self):
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("이력 선택", "삭제할 점검 이력을 먼저 선택하세요.")
            return
        history_iid = selected[0]
        inspection_id = history_iid.replace("history-", "", 1)
        answer = messagebox.askyesno("이력 삭제", "선택한 점검 이력을 삭제할까요?")
        if not answer:
            return
        try:
            delete_inspection_record(int(inspection_id))
            if self.selected_management_no:
                self.refresh_history(self.selected_management_no)
                self.refresh_qr_preview(self.selected_management_no)
            self.status_var.set("선택한 점검 이력을 삭제했습니다.")
        except Exception as exc:
            messagebox.showerror("삭제 실패", str(exc))

    def edit_selected_inspection(self):
        selected = self.history_tree.selection()
        if not selected:
            messagebox.showwarning("이력 선택", "수정할 점검 이력을 먼저 선택하세요.")
            return
        inspection_id = selected[0].replace("history-", "", 1)
        inspection = get_inspection_record(int(inspection_id))
        if not inspection:
            messagebox.showwarning("이력 선택", "수정할 점검 이력을 찾을 수 없습니다.")
            return
        self.open_inspection_dialog(inspection)

    def import_workbook(self):
        initial_dir = str(Path(self.config.get("workbook_path") or Path.home() / "Desktop").parent)
        path = filedialog.askopenfilename(title="검사구 점검계획서 엑셀 선택", initialdir=initial_dir, filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            records = parse_plan_workbook(path)
            created, updated = upsert_tools(records)
            self.config["workbook_path"] = path
            save_config(self.config)
            self.refresh_tool_list()
            self.status_var.set(f"엑셀 반영 완료: 신규 {created}건 / 갱신 {updated}건")
            messagebox.showinfo("엑셀 반영 완료", f"신규 {created}건, 갱신 {updated}건을 반영했습니다.")
        except Exception as exc:
            messagebox.showerror("엑셀 반영 실패", str(exc))

    def export_selected_assets(self):
        if not self.selected_management_no:
            messagebox.showwarning("검사구 선택", "QR을 만들 검사구를 먼저 선택하세요.")
            return
        try:
            qr_path, card_path, _payload = export_tool_assets(self.selected_management_no, self.config)
            self.refresh_qr_preview(self.selected_management_no)
            self.status_var.set(f"QR/HTML 생성 완료: {self.selected_management_no}")
            messagebox.showinfo("생성 완료", f"QR: {qr_path}\nHTML: {card_path}")
        except Exception as exc:
            messagebox.showerror("생성 실패", str(exc))

    def export_all_assets(self):
        rows = list_tools()
        if not rows:
            messagebox.showwarning("데이터 없음", "먼저 검사구 목록을 불러오세요.")
            return
        success = 0
        for row in rows:
            export_tool_assets(row["management_no"], self.config, update_index=False)
            success += 1
        export_index_page(self.config)
        self.status_var.set(f"전체 QR/HTML 갱신 완료: {success}건")
        if self.selected_management_no:
            self.refresh_qr_preview(self.selected_management_no)
        messagebox.showinfo("전체 생성 완료", f"{success}건의 QR/HTML 카드를 갱신했습니다.")

    def open_selected_card(self):
        if not self.selected_management_no:
            messagebox.showwarning("검사구 선택", "열 검사구를 먼저 선택하세요.")
            return
        try:
            _qr_path, card_path, payload = export_tool_assets(self.selected_management_no, self.config, upload_webdav=False)
            if payload.lower().startswith(("http://", "https://")):
                webbrowser.open(payload)
                self.status_var.set(f"웹 카드 열기: {payload}")
            else:
                webbrowser.open(card_path.resolve().as_uri())
                self.status_var.set(f"로컬 카드 열기: {card_path}")
        except Exception as exc:
            messagebox.showerror("카드 열기 실패", str(exc))

    def on_scan_submit(self, _event=None):
        raw = self.scan_var.get().strip()
        if not raw:
            return
        prefix = normalize_text(self.config.get("scan_prefix"))
        management_no = raw
        if prefix and raw.startswith(prefix):
            management_no = raw[len(prefix):].strip()
        elif "/cards/" in raw:
            management_no = Path(raw).stem
        if not get_tool(management_no):
            messagebox.showwarning("조회 실패", f"{management_no} 관리번호를 찾을 수 없습니다.")
            self.status_var.set(f"조회 실패: {management_no}")
            return
        self.search_var.set(management_no)
        self.refresh_tool_list()
        if management_no in self.tool_tree.get_children():
            self.tool_tree.selection_set(management_no)
            self.tool_tree.focus(management_no)
            self.tool_tree.see(management_no)
        self.load_tool(management_no)
        self.scan_var.set("")

    def open_settings(self):
        win = tk.Toplevel(self.root)
        win.title("설정")
        win.geometry("820x470")
        win.configure(bg=UI["panel"])
        vars_map = {
            "workbook_path": tk.StringVar(value=self.config.get("workbook_path", "")),
            "nas_sync_dir": tk.StringVar(value=self.config.get("nas_sync_dir", "")),
            "nas_base_url": tk.StringVar(value=self.config.get("nas_base_url", "")),
            "netlify_site_url": tk.StringVar(value=self.config.get("netlify_site_url", "")),
            "webdav_url": tk.StringVar(value=self.config.get("webdav_url", "")),
            "webdav_username": tk.StringVar(value=self.config.get("webdav_username", "")),
            "webdav_password": tk.StringVar(value=self.config.get("webdav_password", "")),
            "company_name": tk.StringVar(value=self.config.get("company_name", "")),
            "scan_prefix": tk.StringVar(value=self.config.get("scan_prefix", "")),
        }
        for key, label in [
            ("workbook_path", "기본 엑셀 경로"),
            ("nas_sync_dir", "Web Station 폴더"),
            ("nas_base_url", "DDNS / Web URL"),
            ("netlify_site_url", "Netlify URL"),
            ("webdav_url", "WebDAV URL"),
            ("webdav_username", "WebDAV 계정"),
            ("webdav_password", "WebDAV 비밀번호"),
            ("company_name", "기본 담당자"),
            ("scan_prefix", "QR 텍스트 Prefix"),
        ]:
            row = tk.Frame(win, bg=UI["panel"])
            row.pack(fill="x", padx=16, pady=8)
            tk.Label(row, text=label, width=16, anchor="w", bg=UI["panel"], font=("맑은 고딕", 10, "bold")).pack(side="left")
            tk.Entry(row, textvariable=vars_map[key], font=("맑은 고딕", 10)).pack(side="left", fill="x", expand=True)
        guide = (
            "권장 방식: QuickConnect 대신 DDNS + Web Station 주소를 사용하세요.\n"
            "권장 예시: Web Station 폴더=\\\\NAS\\web, DDNS / Web URL=https://sejiqc26.synology.me\n"
            "백슬래시 입력이 어려우면 //192.168.0.2/web 처럼 슬래시로 입력해도 됩니다.\n"
            "cards 폴더를 직접 넣어도 됩니다: \\\\NAS\\web\\cards / https://sejiqc26.synology.me/cards\n"
            "QR은 https://.../cards/SJ-CF-001.html 형식으로 생성됩니다."
        )
        tk.Label(win, text=guide, justify="left", wraplength=660, bg=UI["panel"], fg=UI["muted"], font=("맑은 고딕", 9)).pack(anchor="w", padx=16, pady=(4, 10))

        def save_and_close():
            candidate_url = vars_map["nas_base_url"].get().strip()
            if candidate_url and is_quickconnect_url(candidate_url):
                messagebox.showwarning(
                    "DDNS / Web Station 권장",
                    "QuickConnect 주소는 검사구 QR 공개 링크용으로 적합하지 않습니다.\n"
                    "DDNS 또는 Web Station URL을 넣어 주세요.\n"
                    "예: https://sejiqc26.synology.me/tool-history",
                )
                return
            for key, var in vars_map.items():
                if key == "nas_sync_dir":
                    self.config[key] = normalize_network_path(var.get())
                else:
                    self.config[key] = var.get().strip()
            save_config(self.config)
            self.reset_inspection_form()
            if self.selected_management_no:
                self.refresh_qr_preview(self.selected_management_no)
            self.status_var.set("설정을 저장했습니다.")
            win.destroy()

        self.make_button(win, "저장", save_and_close, "primary", padx=20).pack(anchor="e", padx=16, pady=12)

    def open_settings(self):
        win = tk.Toplevel(self.root)
        win.title("설정")
        win.geometry("840x360")
        win.configure(bg=UI["panel"])

        vars_map = {
            "workbook_path": tk.StringVar(value=self.config.get("workbook_path", "")),
            "netlify_site_url": tk.StringVar(value=self.config.get("netlify_site_url", "")),
            "company_name": tk.StringVar(value=self.config.get("company_name", "")),
            "scan_prefix": tk.StringVar(value=self.config.get("scan_prefix", "")),
        }

        fields = [
            ("workbook_path", "기본 엑셀 경로", ""),
            ("netlify_site_url", "Netlify URL", "예: https://elegant-licorice-178dbe.netlify.app"),
            ("company_name", "기본 담당자", ""),
            ("scan_prefix", "QR 텍스트 Prefix", ""),
        ]

        for key, label, placeholder in fields:
            row = tk.Frame(win, bg=UI["panel"])
            row.pack(fill="x", padx=16, pady=6)
            tk.Label(row, text=label, width=18, anchor="w", bg=UI["panel"], font=("맑은 고딕", 10, "bold")).pack(side="left")
            entry = tk.Entry(row, textvariable=vars_map[key], font=("맑은 고딕", 10))
            entry.pack(side="left", fill="x", expand=True)
            if placeholder:
                tk.Label(row, text=placeholder, width=34, anchor="w", bg=UI["panel"], fg=UI["muted"], font=("맑은 고딕", 9)).pack(side="left", padx=(8, 0))

        guide = (
            "GitHub와 Netlify를 사용하므로 공개 이력카드 주소는 Netlify URL만 사용합니다.\n"
            "설정 저장 후 전체 HTML/QR 갱신을 실행하고 GitHub에 push하면 Netlify가 자동 배포합니다."
        )
        tk.Label(win, text=guide, justify="left", wraplength=780, bg=UI["panel"], fg=UI["muted"], font=("맑은 고딕", 9)).pack(anchor="w", padx=16, pady=(8, 10))

        def save_and_close():
            for key, var in vars_map.items():
                self.config[key] = var.get().strip()
            save_config(self.config)
            self.reset_inspection_form()
            if self.selected_management_no:
                self.refresh_qr_preview(self.selected_management_no)
            self.status_var.set("설정을 저장했습니다.")
            win.destroy()

        self.make_button(win, "저장", save_and_close, "primary", padx=20).pack(anchor="e", padx=16, pady=12)


def main():
    ensure_dirs()
    init_db()
    root = tk.Tk()
    ToolInspectionApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
